import requests
import xml.etree.ElementTree as ET
import pandas as pd
import lxml.etree
import json 
from datetime import timedelta
from datetime import date
import xlwt
import openpyxl
import xlsxwriter

"""***************************************************
Output
***************************************************"""
version = 'v1'
folderName = r'API Data'
currentYear = 2022
currentDate = date.today().strftime('%Y-%m-%d')
reportDate = currentDate
exportFileName = '\Export data - ' + version + ' - ' + reportDate


'''***************************************************
Og link & Var method API
***************************************************'''

#https://api.esv2.com/v2/Api/Messages?apiKey=YOUR_API_KEY_HERE
#https://api.esv2.com/v2/Api/Messages/893?apiKey=YOUR_API_KEY_HERE
#https://api.esv2.com/v2/Api/MessageStatistics/123?apiKey=YOUR_API_KEY_HERE 
#https://api.esv2.com/v2/Api/SmsMmsMessages?apiKey=YOUR_API_KEY_HERE

mtd_eml = 'Api/Messages'
mtd_eml2 = 'Api/Messages/'
mtd_stats = 'Api/MessageStatistics/'
mtd_api = '?apiKey='
mtd_sms = 'Api/SmsMmsMessages/'


'''***************************************************
Test Pandas
***************************************************'''

df_xl = pd.read_excel(r"API Data.xlsx")
#workbook = xlwt.Workbook()
#sheet = workbook.add_sheet('feuille1')
#sheet.write(0, 1, 'cellule en haut a gauche')
#workbook.save('myFile.xls')
#wb = openpyxl.Workbook()
#sheet = wb.active
#c1 = sheet.cell(row = 1, column = 1)
#c1.value = "ANKIT"
#wb.save("demo.xlsx")
workbook = xlsxwriter.Workbook("expertSenderRes.xls")
worksheet = workbook.add_worksheet("first")
worksheet.write(0, 0, "COMPTE")
worksheet.write(0, 1, "SUBJECT")
worksheet.write(0, 2, "ID")
worksheet.write(0, 3, "TYPE")
worksheet.write(0, 4, "SEND DATE")
worksheet.write(0, 5, "FROM NAME")
worksheet.write(0, 6, "FROM EMAIL")
worksheet.write(0, 7, "sent")
worksheet.write(0, 8, "Bounced")

i = 0
for index, row in df_xl.iterrows():
  print('CLE =', row['CLE API'])
  key = row['CLE API']
  r = requests.get(f"{row['SERVEUR']}{mtd_eml}{mtd_api}{row['CLE API']}")
  #print(r.text)
  #r = requests.get('https://api5.esv2.com/v2/Api/Messages?apiKey=XNBsZQVSnDOQfTTFnktE')
  json_data = r.text
  #tree = ET.parse(r.text)
  root = ET.fromstring(json_data)
  for child in root.iter():
    print(' 1 I is = ', i) 
    message = child.find('Id')
    if message is not None:
      message = child.find('Id').text
      compte = child.find('FromEmail').text
      subject = child.find('Subject').text
      type = child.find('Type').text
      sentDate = child.find('SentDate').text
      fromName = child.find('FromName').text
      i += 1
      id_message = message
      msg_stats = requests.get(f'https://api5.esv2.com/v2/Api/MessageStatistics/{id_message}?apiKey={key}')
      each_data = ET.fromstring(msg_stats.text)
      worksheet.write(i, 0, compte)
      worksheet.write(i, 1, subject)
      worksheet.write(i, 2, message)
      worksheet.write(i, 3, type)
      worksheet.write(i, 4, sentDate)
      worksheet.write(i, 5, fromName)
      for c in each_data.iter():
        data = None
        res = None
        t = None
        if c.find('Sent') is not None:
          res = c.find('Sent').text
          worksheet.write(i, 7, res) 
        if c.find('Bounced') is not None:
          t = c.find('Bounced').text
          worksheet.write(i, 8, t)
    if i == 10:
      workbook.close()
   # '''for elem in r.iter():
  #      print (elem)
    
    
    #json_data = r.text
   # root = ET.fromstring(json_data)
  #  textelem = root.find('Message/Id')
  #  print  (textelem)
  #  '''
  
    #dfxml = pd.read_xml(r.content)
    #print (dfxml)
    #root = ET.fromstring(json)
    #ET.dump(root)
    

'''OLD TEST'''

#root = lxml.etree.fromstring(xmlstr)
#textelem = root.find('result/field/value/text')
#print textelem.text
    #data =[]
    #testid = root.find('Id')
    #for id in root.iter('Message'):
        #print (id.text)
        #for child in id.findall('Id'):
            #ID = child.text
            #print (ID)
            #v1 = dict(tag = child.tag, text = child.text)



#MsgId = []

#for child in root.iter('Id'):
    #print(child.tag, child.text)

#SentDate = []

#for child in root.iter('SentDate'):
 #   SentDate.append(child.text)